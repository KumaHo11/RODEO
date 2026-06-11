from openpyxl import load_workbook

wb = load_workbook('sheet.xlsx', data_only=False)
ws = wb.active

header_row = 1
for r in range(1, 10):
    for c in range(1, 300):
        if ws.cell(row=r, column=c).value == 'Mínimo dias pastoreo':
            header_row = r
            break

print(f"Header row: {header_row}")

for r in range(header_row + 1, header_row + 5):
    for c in range(1, 300):
        val = ws.cell(row=header_row, column=c).value
        if val in ['Mínimo dias pastoreo', 'Maximo dias pastoreo', 'Mínimo dias pastoreo ', 'Maximo dias pastoreo ']:
            formula = ws.cell(row=r, column=c).value
            print(f"Row {r}, Col {c} ({val}): {formula}")
