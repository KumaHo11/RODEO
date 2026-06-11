import openpyxl

wb = openpyxl.load_workbook('/Users/javi/RODEO/sheet.xlsx', data_only=False)
sheet = wb.active

for col in range(1, 10):
    header = sheet.cell(row=8, column=col).value
    val = sheet.cell(row=10, column=col).value
    formula = sheet.cell(row=10, column=col).value
    print(f"Col {col}: Header='{header}', Value='{val}', Formula='{formula}'")
