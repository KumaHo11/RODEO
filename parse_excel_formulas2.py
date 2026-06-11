from openpyxl import load_workbook

wb = load_workbook('sheet.xlsx', data_only=True)
ws = wb.active

print(f"G60: {ws['G60'].value}")
print(f"W60: {ws['W60'].value}")
print(f"B53: {ws['B53'].value}")
print(f"G58: {ws['G58'].value} ({ws['F58'].value})")
print(f"G59: {ws['G59'].value} ({ws['F59'].value})")
print(f"HW59: {ws['HW59'].value}")

