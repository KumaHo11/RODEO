from openpyxl import load_workbook

wb = load_workbook('sheet.xlsx', data_only=False)
ws = wb.active

print(f"Row 17 Min: {ws.cell(row=17, column=225).value} | {ws.cell(row=17, column=227).value}")
print(f"B17: {ws['B17'].value}")
print(f"B53: {ws['B53'].value}")
print(f"W60: {ws['W60'].value} (Formula: {ws['W60'].value if isinstance(ws['W60'].value, str) else 'Literal'})")
print(f"G60: {ws['G60'].value}")
