from openpyxl import load_workbook

wb = load_workbook('sheet.xlsx', data_only=False)
ws = wb.active

for r in range(9, 20):
    val = ws.cell(row=r, column=226).value # Maximo dias pastoreo
    val2 = ws.cell(row=r, column=228).value # Maximo dias pastoreo
    print(f"Row {r} Max: {val} | {val2}")
