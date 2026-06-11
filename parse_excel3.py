import openpyxl

wb = openpyxl.load_workbook('/Users/javi/RODEO/sheet.xlsx', data_only=False)
sheet = wb.active

cells_to_check = ['B10', 'G60', 'AD70', 'BI70', 'IA5', 'HW10']

for cell in cells_to_check:
    print(f"Cell {cell} formula: {sheet[cell].value}")

# Search for "Total" in the entire sheet
for r in range(1, 100):
    for c in range(200, 260): # HN is around column 222
        val = sheet.cell(row=r, column=c).value
        if val and isinstance(val, str) and 'E: Total de días de pastoreo requeridos' in val:
            print(f"Found 'E: Total de días...' at Row {r} Col {c}")
            # check the next few columns for formula
            for rc in range(c, c+15):
                f = sheet.cell(row=r, column=rc).value
                if f:
                    print(f"Value/Formula at Row {r} Col {rc}: {f}")
