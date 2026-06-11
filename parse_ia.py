from openpyxl import load_workbook

wb = load_workbook('sheet.xlsx', data_only=False)
ws = wb.active

print(f"IA10 Formula: {ws['IA10'].value}")
