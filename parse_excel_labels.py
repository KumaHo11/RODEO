from openpyxl import load_workbook

wb = load_workbook('sheet.xlsx', data_only=True)
ws = wb.active

for r in range(50, 65):
    for c in ['E', 'F', 'G', 'U', 'V', 'W', 'X']:
        val = ws[f"{c}{r}"].value
        if val is not None:
            print(f"{c}{r}: {val}")
