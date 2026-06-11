from openpyxl import load_workbook

wb = load_workbook('sheet.xlsx', data_only=True)
ws = wb.active

print(f"IA10: {ws['IA10'].value}")
print(f"HW10: {ws['HW10'].value}")
print(f"HZ10: {ws['HZ10'].value}")
print(f"IB10: {ws['IB10'].value}")
