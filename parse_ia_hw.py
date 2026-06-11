from openpyxl import load_workbook

wb = load_workbook('sheet.xlsx', data_only=True)
ws = wb.active

print(f"IA8: {ws['IA8'].value}")
print(f"HW8: {ws['HW8'].value}")
