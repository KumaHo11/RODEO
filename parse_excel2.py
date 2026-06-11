import openpyxl

wb = openpyxl.load_workbook('/Users/javi/RODEO/sheet.xlsx', data_only=False)
sheet = wb.active

cols_of_interest = [225, 226, 227, 228, 233, 234, 235, 236]

print("Formulas in Row 10:")
for c in cols_of_interest:
    header = sheet.cell(row=8, column=c).value
    formula = sheet.cell(row=10, column=c).value
    print(f"Col {openpyxl.utils.get_column_letter(c)} ({header}): {formula}")

# Also check for "Total de dias de pastoreo requerido"
# In the CSV it was around row 59: "E: Total de días de pastoreo requeridos,,,,,,,112,,,,,"
# Let's search row 59 for "Total de días"
for r in range(50, 65):
    for c in range(1, 20):
        val = str(sheet.cell(row=r, column=c).value)
        if 'Total de días de pastoreo' in val or 'Total de dias de pastoreo' in val:
            # Check the cells to the right for a formula
            for right_c in range(c, c+15):
                f = sheet.cell(row=r, column=right_c).value
                if f and isinstance(f, str) and f.startswith('='):
                    print(f"Total dias pastoreo at Row {r} Col {right_c}: {f}")
                elif f and isinstance(f, (int, float)):
                    print(f"Total dias pastoreo value at Row {r} Col {right_c}: {f}")
