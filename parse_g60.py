from openpyxl import load_workbook

wb = load_workbook('sheet.xlsx', data_only=False)
ws = wb.active

print(f"HW59 ({ws['HW58'].value}): {ws['HW59'].value}")
print(f"G58 ({ws['F58'].value}): {ws['G58'].value}")
print(f"G59 ({ws['F59'].value}): {ws['G59'].value}")
