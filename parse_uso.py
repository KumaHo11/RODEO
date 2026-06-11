from openpyxl import load_workbook

wb = load_workbook('sheet.xlsx', data_only=False)
ws = wb.active

for r in range(9, 15):
    val = ws.cell(row=r, column=231).value # % USO is usually around there, let's search it
    for c in range(220, 240):
        if ws.cell(row=8, column=c).value == '%USO':
            print(f"Row {r} %USO (Col {c}): {ws.cell(row=r, column=c).value}")
