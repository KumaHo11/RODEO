import openpyxl
import sys

# Load the workbook
wb = openpyxl.load_workbook('/Users/javi/RODEO/sheet.xlsx', data_only=False)
sheet = wb.active # Get active sheet

# Find column names from row 8
for col in range(1, sheet.max_column + 1):
    val = sheet.cell(row=8, column=col).value
    if val:
        val_str = str(val).strip()
        if any(x in val_str.lower() for x in ['mínimo', 'maximo', 'uso', 'rac', 'rend', 'dias de pastoreo']):
            print(f"Col {col} ({openpyxl.utils.get_column_letter(col)}): {val_str}")

# Print formula for a sample row (Row 10) for these specific columns
print("\nFormulas for row 10:")
for col_letter in ['IW', 'IX', 'IY', 'IZ', 'JA', 'JB', 'JC', 'JD', 'JE', 'JF', 'JG', 'JH', 'JI']:
    try:
        col = openpyxl.utils.column_index_from_string(col_letter)
        header = sheet.cell(row=8, column=col).value
        formula = sheet.cell(row=10, column=col).value
        val = sheet.cell(row=10, column=col).value
        print(f"Col {col_letter} ({header}): {formula}")
    except Exception as e:
        pass

# Also let's check row 59 for Total de dias de pastoreo
total_dias = sheet.cell(row=59, column=openpyxl.utils.column_index_from_string('N')).value
print(f"\nRow 59, Col N (Total de días de pastoreo requeridos): {total_dias}")
